from openpyxl import load_workbook
import pandas as pd


def get_email_rows(excel_file) -> pd.DataFrame:
    wb = load_workbook(excel_file, data_only=True)
    sh = wb['Sheet1']

    df = pd.DataFrame(columns=['type', 'name', 'contact'])

    for line in sh.iter_rows():
        if any(map(lambda x: True if 'FF' in str(x.font.color.index) else False, line)) or line[0].value is not None:
            data = list(line)
            actual_data = [x.value for x in data if x is not None]
            df.loc[0 if pd.isnull(df.index.max()) else df.index.max() + 1] = actual_data

    df.type = df.type.fillna(method='pad')
    df = df.dropna(how='any')
    return df
